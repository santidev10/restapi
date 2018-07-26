import logging
from os import listdir
from os.path import isfile, join

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management import BaseCommand
from openpyxl import load_workbook

from segment.models import SegmentKeyword, SegmentVideo, SegmentChannel

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    fixtures_directory = join(settings.BASE_DIR, "segment/fixtures/")
    file_name_prefix = "segments_private_export_"
    separation_symbol = "|"
    fields_with_separation_symbol = ["related_ids", "shared_with"]
    names_and_models = {
        "KeywordSegments": SegmentKeyword,
        "VideoSegments": SegmentVideo,
        "ChannelSegments": SegmentChannel
    }

    def handle(self, *args, **options):
        logger.info("Start importing viewiq users")
        workbooks = self.__load_workbooks()
        parsed_data = self.__parse_workbooks(workbooks)
        self.__create_segments(parsed_data)
        logger.info("Import viewiq users procedure has been finished")

    def __load_workbooks(self):
        import_files = [
            file_name for file_name in listdir(self.fixtures_directory)
            if isfile(join(self.fixtures_directory, file_name))
            and file_name.startswith(self.file_name_prefix)
        ]
        workbooks = [load_workbook(filename=join(self.fixtures_directory, fixture))
                     for fixture in import_files]
        return workbooks

    def __parse_workbooks(self, workbooks):
        result = {}
        for workbook in workbooks:
            for worksheet in workbook.worksheets:
                data = worksheet.values
                headers = next(data)
                headers = [header.lower().replace(" ", "_") for header in headers]
                parsed_data = [dict(zip(headers, obj)) for obj in data]
                parsed_data = self.__clean_up_separation_symbols(parsed_data)
                result[worksheet.title] = parsed_data
        return result

    def __clean_up_separation_symbols(self, data):
        for obj in data:
            for key in self.fields_with_separation_symbol:
                if obj[key] is not None:
                    obj[key] = obj[key].split(self.separation_symbol)
        return data

    def __create_segments(self, segments_data):
        for key, value in segments_data.items():
            segment_model = self.names_and_models[key]
            for obj in value:
                obj["owner"] = None
                owner_email = obj.pop("owner_email")
                related_ids = obj.pop("related_ids")
                if obj.get("shared_with") is None:
                    obj["shared_with"] = []
                if owner_email:
                    try:
                        obj["owner"] = get_user_model().objects.get(email=owner_email)
                    except get_user_model().DoesNotExist:
                        pass
                segment = segment_model.objects.create(**obj)
                if related_ids:
                    segment.add_related_ids(related_ids)
                segment.update_statistics(segment)
