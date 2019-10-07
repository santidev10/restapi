from datetime import date
from datetime import timedelta
from io import BytesIO
from unittest import skip
from unittest.mock import patch

from django.db.models.signals import post_save
from django.test import TransactionTestCase
from openpyxl import load_workbook

from ads_analyzer.models import OpportunityTargetingReport
from ads_analyzer.models.opportunity_targeting_report import ReportStatus
from ads_analyzer.reports.opportunity_targeting_report.s3_exporter import OpportunityTargetingReportS3Exporter
from ads_analyzer.reports.opportunity_targeting_report.serializers import TargetTableSerializer
from ads_analyzer.tasks import create_opportunity_targeting_report
from aw_reporting.models import AdGroup
from aw_reporting.models import Campaign
from aw_reporting.models import KeywordStatistic
from aw_reporting.models import OpPlacement
from aw_reporting.models import Opportunity
from aw_reporting.models import Topic
from aw_reporting.models import TopicStatistic
from aw_reporting.models import YTChannelStatistic
from aw_reporting.models import YTVideoStatistic
from aw_reporting.models.salesforce_constants import SalesForceGoalType
from email_reports.tasks import notify_opportunity_targeting_report_is_ready
from saas import celery_app
from utils.utittests.celery import mock_send_task
from utils.utittests.patch_now import patch_now
from utils.utittests.s3_mock import S3TestCase
from utils.utittests.str_iterator import str_iterator


class CreateOpportunityTargetingReportBaseTestCase(TransactionTestCase, S3TestCase):
    def act(self, opportunity_id, date_from, date_to):
        create_opportunity_targeting_report(
            opportunity_id=opportunity_id,
            date_from_str=str(date_from),
            date_to_str=str(date_to),
        )

    def get_report_workbook(self, opportunity_id, date_from, date_to):
        s3_key = OpportunityTargetingReportS3Exporter.get_s3_key(
            opportunity_id,
            str(date_from),
            str(date_to)
        )
        self.assertTrue(OpportunityTargetingReportS3Exporter.exists(s3_key, get_key=False))
        file = OpportunityTargetingReportS3Exporter.get_s3_export_content(s3_key, get_key=False)
        book = load_workbook(BytesIO(file.read()))
        return book


class CreateOpportunityTargetingReportGeneralTestCase(CreateOpportunityTargetingReportBaseTestCase):
    def test_empty_report_updates_db_entity(self):
        opportunity = Opportunity.objects.create(id=next(str_iterator))
        date_from, date_to = date(2020, 1, 1), date(2020, 1, 2)
        with patch.object(post_save, "send"):
            report = OpportunityTargetingReport.objects.create(
                opportunity=opportunity,
                date_from=date_from,
                date_to=date_to,
            )
        self.act(opportunity.id, date_from, date_to)

        report.refresh_from_db()
        s3_key = OpportunityTargetingReportS3Exporter.get_s3_key(opportunity.id, str(date_from), str(date_to))
        self.assertEqual(s3_key, report.s3_file_key)
        self.assertEqual(ReportStatus.SUCCESS.value, report.status)

    def test_send_email_notifications(self):
        opportunity = Opportunity.objects.create(id=next(str_iterator))
        date_from, date_to = date(2020, 1, 1), date(2020, 1, 2)

        with mock_send_task():
            self.act(opportunity.id, date_from, date_to)

            calls = celery_app.send_task.mock_calls

        self.assertEqual(1, len(calls))
        expected_kwargs = dict(
            opportunity_id=opportunity.id,
            date_from_str=str(date_from),
            date_to_str=str(date_to),
        )
        self.assertEqual(
            (notify_opportunity_targeting_report_is_ready.name, (), expected_kwargs),
            calls[0][1]
        )

    def test_empty_report_to_s3(self):
        opportunity = Opportunity.objects.create(id=next(str_iterator))
        date_from, date_to = date(2020, 1, 1), date(2020, 1, 2)

        self.act(opportunity.id, date_from, date_to)

        book = self.get_report_workbook(opportunity.id, date_from, date_to)
        self.assertIsNotNone(book)

    def test_empty_report_header(self):
        opportunity = Opportunity.objects.create(
            id=next(str_iterator),
            name="Test Opportunity name",
        )
        date_from, date_to = date(2020, 1, 1), date(2020, 1, 2)

        self.act(opportunity.id, date_from, date_to)

        book = self.get_report_workbook(opportunity.id, date_from, date_to)
        self.assertEqual(
            ["Target", "Devices", "Demo", "Video"],
            book.sheetnames
        )
        for sheet, sheet_name in zip(book.worksheets, book.sheetnames):
            with self.subTest(sheet_name):
                self.assertEqual(
                    f"Opportunity: {opportunity.name}",
                    sheet.cell(None, 1, 1).value,
                )
                self.assertEqual(
                    f"Date Range: {date_from} - {date_to}",
                    sheet.cell(None, 2, 1).value,
                )


class ColumnsDeclaration:
    def __init__(self, definition):
        self.definition = definition
        self.values = dict(definition)

    def __getattr__(self, item):
        try:
            return super().__getattribute__(item)
        except AttributeError:
            return self.values[item]

    def labels(self):
        return tuple([name for _, name in self.definition])

    def ids(self):
        return tuple([name for name, _ in self.definition])


class CreateOpportunityTargetingReportSheetTestCase(CreateOpportunityTargetingReportBaseTestCase):
    SHEET_NAME = None
    DATA_ROW_INDEX = 3
    columns: ColumnsDeclaration = None

    def get_data_table(self, opportunity_id, date_from, date_to):
        book = self.get_report_workbook(opportunity_id, date_from, date_to)
        sheet = book.get_sheet_by_name(self.SHEET_NAME)
        return list(sheet)[self.DATA_ROW_INDEX:]

    def setUp(self) -> None:
        super().setUp()
        self.opportunity = Opportunity.objects.create(
            id=next(str_iterator),
            name="Test opportunity #123",
        )


class CreateOpportunityTargetingReportTargetTestCase(CreateOpportunityTargetingReportSheetTestCase):
    SHEET_NAME = "Target"

    columns = ColumnsDeclaration(
        (
            ("target", "Target"),
            ("type", "Type"),
            ("campaign_name", "Ads Campaign"),
            ("ad_group_name", "Ads Ad group"),
            ("placement_name", "Salesforce Placement"),
            ("placement_start", "Placement Start Date"),
            ("placement_end", "Placement End Date"),
            ("days_remaining", "Days remaining"),
            ("margin_cap", "Margin Cap"),
            ("cannot_roll_over", "Cannot Roll over Delivery"),
            ("goal_type", "Rate Type"),
            ("contracted_rate", "Contracted Rate"),
            ("max_bid", "Max bid"),
            ("avg_rate", "Avg. Rate"),
            ("cost", "Cost"),
            ("cost_delivered_percentage", "Cost delivery percentage"),
            ("impressions", "Impressions"),
            ("views", "Views"),
            ("delivery_percentage", "Delivery percentage"),
            ("revenue", "Revenue"),
            ("profit", "Profit"),
            ("margin", "Margin"),
            ("video_played_100", "Video played to 100%"),
            ("view_rate", "View rate"),
            ("clicks", "Clicks"),
            ("ctr", "CTR"),
        )
    )

    def setUp(self) -> None:
        super().setUp()
        any_date = date(2019, 1, 1)
        pl_start, pl_end = any_date - timedelta(days=1), any_date + timedelta(days=1)
        self.placement = OpPlacement.objects.create(opportunity=self.opportunity, name="Test Placement",
                                                    goal_type_id=SalesForceGoalType.CPV, ordered_rate=1.23,
                                                    start=pl_start, end=pl_end)
        self.campaign = Campaign.objects.create(salesforce_placement=self.placement, name="Test Campaign")
        self.ad_group = AdGroup.objects.create(campaign=self.campaign, name="Test AdGroup")


class CreateOpportunityTargetingReportTargetDataTestCase(CreateOpportunityTargetingReportTargetTestCase):
    def get_data_dict(self, *args, **kwargs):
        rows = self.get_data_table(*args, **kwargs)
        headers = [cell.value for cell in rows[0]]

        return [
            dict(zip(headers, [cell.value for cell in row]))
            for row in rows[1:]
        ]

    def test_headers(self):
        any_date = date(2019, 1, 1)
        self.act(self.opportunity.id, any_date, any_date)

        rows = self.get_data_table(self.opportunity.id, any_date, any_date)
        headers = rows[0]
        self.assertEqual(
            list(self.columns.labels()),
            [cell.value for cell in headers]
        )

    def test_general_data(self):
        self.opportunity.cannot_roll_over = True
        self.opportunity.save()
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(self.campaign.name, item[columns.campaign_name])
        self.assertEqual(self.ad_group.name, item[columns.ad_group_name])
        self.assertEqual(self.placement.name, item[columns.placement_name])
        self.assertEqual(str(self.placement.start), item[columns.placement_start])
        self.assertEqual(str(self.placement.end), item[columns.placement_end])
        self.assertEqual(self.opportunity.cannot_roll_over, item[columns.cannot_roll_over])
        self.assertEqual(self.placement.goal_type, item[columns.goal_type])
        self.assertEqual(self.placement.ordered_rate, item[columns.contracted_rate])

    def test_topic_general_data(self):
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(topic.name, item[columns.target])
        self.assertEqual("Topic", item[columns.type])

    def test_keyword_general_data(self):
        any_date = date(2019, 1, 1)
        self.opportunity.cannot_roll_over = True
        self.opportunity.save()
        keyword = "test keyword"
        KeywordStatistic.objects.create(keyword=keyword, ad_group=self.ad_group, date=any_date)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(keyword, item[columns.target])
        self.assertEqual("Keyword", item[columns.type])

    @skip("Not implemented")
    def test_interests_in_marketing_general_data(self):
        raise NotImplemented

    @skip("Not implemented")
    def test_interests_affinity_general_data(self):
        raise NotImplemented

    @skip("Not implemented")
    def test_interests_caa_general_data(self):
        raise NotImplemented

    @skip("Not implemented")
    def test_interests_custom_intent_general_data(self):
        raise NotImplemented

    @skip("Not implemented")
    def test_interests_detailed_demographic_general_data(self):
        raise NotImplemented

    @skip("Not implemented")
    def test_channel_general_data(self):
        any_date = date(2019, 1, 1)
        channel_id = next(str_iterator)
        YTChannelStatistic.objects.create(yt_id=channel_id, ad_group=self.ad_group)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(channel_id, item[columns.target])
        self.assertEqual("Channel", item[columns.type])

    @skip("Not implemented")
    def test_channel_general_data_title(self):
        raise NotImplementedError

    @skip("Not implemented")
    def test_video_general_data(self):
        any_date = date(2019, 1, 1)
        video_id = next(str_iterator)
        YTVideoStatistic.objects.create(yt_id=video_id, ad_group=self.ad_group)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(video_id, item[columns.target])
        self.assertEqual("Video", item[columns.type])

    @skip("Not implemented")
    def test_video_general_data_title(self):
        raise NotImplementedError

    def test_general_stats(self):
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              impressions=1000, video_views=200, cost=1.02, clicks=30)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(stats.impressions, item[columns.impressions])
        self.assertEqual(stats.video_views, item[columns.views])
        self.assertEqual(stats.cost, item[columns.cost])
        self.assertEqual(stats.clicks, item[columns.clicks])

    def test_days_remaining(self):
        any_date = date(2019, 1, 1)
        days_remaining = 3
        test_now = self.placement.end - timedelta(days=days_remaining)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch_now(test_now):
            self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(days_remaining, item[columns.days_remaining])

    @skip("Not implemented")
    def test_margin_cap(self):
        raise NotImplementedError

    @skip("Not implemented")
    def test_max_bid(self):
        raise NotImplementedError

    def test_average_rate_cpv(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPV
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              cost=23, video_views=34)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        expected_rate = stats.cost / stats.video_views
        self.assertAlmostEqual(expected_rate, item[columns.avg_rate])

    def test_average_rate_cpm(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPM
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              cost=23, impressions=34)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        expected_rate = stats.cost / (stats.impressions / 1000)
        self.assertAlmostEqual(expected_rate, item[columns.avg_rate])

    def test_cost_delivery_percentage(self):
        any_date = date(2019, 1, 1)
        topic_1 = Topic.objects.create(name="Test topic 1")
        topic_2 = Topic.objects.create(name="Test topic 2")
        costs = (20, 80)
        stats_1 = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic_1,
                                                date=any_date,
                                                cost=costs[0])
        stats_2 = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic_2,
                                                date=any_date,
                                                cost=costs[1])

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(2, len(data))
        columns = self.columns
        topics_cost_delivery_percentage = {
            item[columns.target]: item[columns.cost_delivered_percentage]
            for item in data
        }
        sum_cost = sum(costs)
        self.assertEqual(
            stats_1.cost / sum_cost,
            topics_cost_delivery_percentage[topic_1.name]
        )
        self.assertEqual(
            stats_2.cost / sum_cost,
            topics_cost_delivery_percentage[topic_2.name]
        )

    def test_delivery_percentage_cpm(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPM
        self.placement.save()
        topic_1 = Topic.objects.create(name="Test topic 1")
        topic_2 = Topic.objects.create(name="Test topic 2")
        impressions = (20, 80)
        stats_1 = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic_1,
                                                date=any_date,
                                                impressions=impressions[0])
        stats_2 = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic_2,
                                                date=any_date,
                                                impressions=impressions[1])

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(2, len(data))
        columns = self.columns
        topics_cost_delivery_percentage = {
            item[columns.target]: item[columns.delivery_percentage]
            for item in data
        }
        sum_impressions = sum(impressions)
        self.assertEqual(
            stats_1.impressions / sum_impressions,
            topics_cost_delivery_percentage[topic_1.name]
        )
        self.assertEqual(
            stats_2.impressions / sum_impressions,
            topics_cost_delivery_percentage[topic_2.name]
        )

    def test_delivery_percentage_cpv(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPV
        self.placement.save()
        topic_1 = Topic.objects.create(name="Test topic 1")
        topic_2 = Topic.objects.create(name="Test topic 2")
        views = (20, 80)
        stats_1 = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic_1,
                                                date=any_date,
                                                video_views=views[0])
        stats_2 = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic_2,
                                                date=any_date,
                                                video_views=views[1])

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(2, len(data))
        columns = self.columns
        topics_cost_delivery_percentage = {
            item[columns.target]: item[columns.delivery_percentage]
            for item in data
        }
        sum_views = sum(views)
        self.assertEqual(
            stats_1.video_views / sum_views,
            topics_cost_delivery_percentage[topic_1.name]
        )
        self.assertEqual(
            stats_2.video_views / sum_views,
            topics_cost_delivery_percentage[topic_2.name]
        )

    def test_revenue_cpv(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPV
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              video_views=34)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertAlmostEqual(
            stats.video_views * self.placement.ordered_rate,
            item[columns.revenue]
        )

    def test_revenue_cpm(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPM
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              impressions=3400)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertAlmostEqual(
            stats.impressions * self.placement.ordered_rate / 1000,
            item[columns.revenue]
        )

    def test_profit(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPV
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              video_views=34, impressions=450, cost=13)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        revenue = stats.video_views * self.placement.ordered_rate
        expected_profit = revenue - stats.cost
        self.assertAlmostEqual(
            expected_profit,
            item[columns.profit]
        )

    def test_margin(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPV
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              video_views=34, impressions=450, cost=13)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        revenue = stats.video_views * self.placement.ordered_rate
        profit = revenue - stats.cost
        expected_margin = profit / revenue
        self.assertAlmostEqual(
            expected_margin,
            item[columns.margin]
        )

    def test_video_played_to_100(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPV
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              video_views_100_quartile=34.2, impressions=450)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        expected_video_played_100 = stats.video_views_100_quartile / stats.impressions
        self.assertAlmostEqual(
            expected_video_played_100,
            item[columns.video_played_100]
        )

    def test_view_rate_cpv(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPV
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              impressions=200, video_views=30)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertAlmostEqual(stats.video_views / stats.impressions, item[columns.view_rate])

    def test_view_rate_cpm(self):
        any_date = date(2019, 1, 1)
        self.placement.goal_type_id = SalesForceGoalType.CPM
        self.placement.save()
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                      impressions=200, clicks=30)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(None, item[columns.view_rate])

    def test_ctr(self):
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        stats = TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date,
                                              impressions=200, clicks=30)

        self.act(self.opportunity.id, any_date, any_date)
        data = self.get_data_dict(self.opportunity.id, any_date, any_date)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertAlmostEqual(stats.clicks / stats.impressions, item[columns.ctr])

    @skip("Not implemented")
    def test_ordering(self):
        raise NotImplementedError

    def test_general_stats_aggregates(self):
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        impressions = (1000, 2100)
        period = len(impressions)
        date_from, date_to = any_date, any_date + timedelta(days=period)
        for index in range(period):
            TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic,
                                          date=any_date + timedelta(days=index),
                                          impressions=impressions[index])

        self.act(self.opportunity.id, date_from, date_to)
        data = self.get_data_dict(self.opportunity.id, date_from, date_to)
        self.assertEqual(1, len(data))
        item = data[0]
        columns = self.columns
        self.assertEqual(sum(impressions), item[columns.impressions])


class CreateOpportunityTargetingReportTargetFormattingTestCase(CreateOpportunityTargetingReportTargetTestCase):
    class Color:
        RED = "FFFF0013"
        YELLOW = "FFFFFE50"
        GREEN = "FF00B25B"

    def get_cell_dict(self, *args, **kwargs):
        rows = self.get_data_table(*args, **kwargs)
        headers = [cell.value for cell in rows[0]]

        return [
            dict(zip(headers, row))
            for row in rows[1:]
        ]

    def test_cost_delivery_percentage_red(self):
        test_cost_delivered_percentage = .45
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_cost_delivery_percentage",
                          return_value=test_cost_delivered_percentage):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.RED, item[columns.cost_delivered_percentage].fill.start_color.rgb)

    def test_cost_delivery_percentage_yellow(self):
        test_cost_delivered_percentage = .55
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_cost_delivery_percentage",
                          return_value=test_cost_delivered_percentage):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.YELLOW, item[columns.cost_delivered_percentage].fill.start_color.rgb)

    def test_cost_delivery_percentage_green(self):
        test_cost_delivered_percentage = .75
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_cost_delivery_percentage",
                          return_value=test_cost_delivered_percentage):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.GREEN, item[columns.cost_delivered_percentage].fill.start_color.rgb)

    def test_delivery_percentage_red(self):
        test_delivered_percentage = .45
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_delivery_percentage", return_value=test_delivered_percentage):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.RED, item[columns.delivery_percentage].fill.start_color.rgb)

    def test_delivery_percentage_yellow(self):
        test_delivered_percentage = .55
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_delivery_percentage", return_value=test_delivered_percentage):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.YELLOW, item[columns.delivery_percentage].fill.start_color.rgb)

    def test_delivery_percentage_green(self):
        test_delivered_percentage = .75
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_delivery_percentage", return_value=test_delivered_percentage):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.GREEN, item[columns.delivery_percentage].fill.start_color.rgb)

    def test_margin_red(self):
        test_margin = .25
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_margin", return_value=test_margin):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.RED, item[columns.margin].fill.start_color.rgb)

    def test_margin_yellow(self):
        test_margin = .35
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_margin", return_value=test_margin):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.YELLOW, item[columns.margin].fill.start_color.rgb)

    def test_margin_green(self):
        test_margin = .45
        any_date = date(2019, 1, 1)
        topic = Topic.objects.create(name="Test topic")
        TopicStatistic.objects.create(ad_group=self.ad_group, topic=topic, date=any_date)

        with patch.object(TargetTableSerializer, "get_margin", return_value=test_margin):
            self.act(self.opportunity.id, any_date, any_date)
        styles = self.get_cell_dict(self.opportunity.id, any_date, any_date)
        item = styles[0]
        columns = self.columns
        self.assertEqual(self.Color.GREEN, item[columns.margin].fill.start_color.rgb)

    @skip("Not implemented")
    def test_percentage_fields(self):
        raise NotImplementedError
