import hashlib
import logging
from datetime import datetime
from os import listdir
from os.path import isfile, join

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management import BaseCommand
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from userprofile.models import UserChannel

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    fixtures_directory = join(settings.BASE_DIR, "userprofile/fixtures/")
    file_name_prefix = "users_export_"
    separation_symbol = "|"
    fields_with_separation_symbol = ["related_channels", "access"]
    default_google_id = "Yes, this is a google id"
    admin_user_email = "admin@admin.admin"

    def handle(self, *args, **options):
        logger.info("Start importing viewiq users")
        workbooks = self.__load_workbooks()
        parsed_data = self.__parse_workbooks(workbooks)
        self.__create_users(parsed_data)
        logger.info("Import viewiq users procedure has been finished")

    def __load_workbooks(self):
        import_files = [
            file_name for file_name in listdir(self.fixtures_directory)
            if isfile(join(self.fixtures_directory, file_name))
            and file_name.startswith(self.file_name_prefix)
        ]
        workbooks = [load_workbook(filename=join(self.fixtures_directory, fixture))
                     for fixture in import_files]
        return workbooks

    def __parse_workbooks(self, workbooks):
        result = []
        for workbook in workbooks:
            for worksheet in workbook.worksheets:
                data = worksheet.values
                headers = next(data)
                headers = [header.lower().replace(" ", "_") for header in headers]
                parsed_data = [dict(zip(headers, obj)) for obj in data]
                parsed_data = self.__clean_up_separation_symbols(parsed_data)
                result += parsed_data
        return result

    def __clean_up_separation_symbols(self, data):
        for obj in data:
            for key in self.fields_with_separation_symbol:
                if obj[key] is not None:
                    obj[key] = obj[key].split(self.separation_symbol)
        return data

    @transaction.atomic
    def __create_users(self, users_data):
        for user_data in users_data:
            email = user_data.get("email")
            if email == self.admin_user_email:
                continue
            related_channels = user_data.pop("related_channels") or []
            access = user_data.pop("access") or []
            date_joined = datetime.strptime(user_data["date_joined"], "%Y-%m-%d %H:%M:%S")
            date_joined = date_joined.replace(tzinfo=timezone.utc)
            company = user_data.get("company")
            phone_number = user_data.get("phone_number")
            try:
                user = get_user_model().objects.get(email=email)
            except get_user_model().DoesNotExist:
                user_data["password"] = hashlib.sha1(str(timezone.now().timestamp()).encode()).hexdigest()
                user_data["first_name"] = user_data["first_name"] or ""
                user_data["last_name"] = user_data["first_name"] or ""
                last_login = datetime.strptime(user_data["last_login"], "%Y-%m-%d %H:%M:%S")
                last_login = last_login.replace(tzinfo=timezone.utc)
                user_data["last_login"] = last_login
                user_data["date_joined"] = date_joined
                user = get_user_model().objects.create(**user_data)
                user.set_password(user.password)
                for obj in access:
                    user.add_custom_user_group(obj)
            else:
                user.date_joined = min(date_joined, user.date_joined)
                if company and not user.company:
                    user.company = company
                if phone_number and not user.phone_number:
                    user.phone_number = phone_number
            user.google_account_id = self.default_google_id
            user.save()
            for channel_id in related_channels:
                UserChannel.objects.get_or_create(user=user, channel_id=channel_id)
