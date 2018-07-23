import io
import json
import re
from itertools import product

from django.core.urlresolvers import reverse
from django.http import HttpResponse
from openpyxl import load_workbook
from rest_framework.status import HTTP_200_OK

from aw_creation.api.urls.names import Name
from aw_creation.models import AccountCreation
from aw_reporting.api.constants import DashboardRequest
from aw_reporting.demo.models import DEMO_ACCOUNT_ID
from aw_reporting.excel_reports import FOOTER_ANNOTATION
from aw_reporting.models import Account, Campaign
from saas.urls.namespaces import Namespace
from utils.utils_tests import ExtendedAPITestCase, int_iterator


class AnalyzeExportAPITestCase(ExtendedAPITestCase):
    def _get_url(self, account_creation_id):
        return reverse(
            Namespace.AW_CREATION + ":" + Name.Dashboard.PERFORMANCE_EXPORT_WEEKLY_REPORT,
            args=(account_creation_id,))


class AnalyzeExportAnalyticsAPITestCase(AnalyzeExportAPITestCase):
    def test_success(self):
        user = self.create_test_user()
        account = Account.objects.create(id=1, name="")
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account)

        url = self._get_url(account_creation.id)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)
        self.assertEqual(type(response), HttpResponse)

    def test_success_demo(self):
        self.create_test_user()
        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)
        self.assertEqual(type(response), HttpResponse)

    def test_quartiles_are_percent_formatted(self):
        self.create_test_user()
        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        f = io.BytesIO(response.content)
        book = load_workbook(f)
        sheet = book.worksheets[0]

        row_indexes = range(14, 17)
        column_indexes = range(7, 11)
        cell_indexes = product(row_indexes, column_indexes)
        for row, column in cell_indexes:
            cell = sheet[row][column]
            self.assertEqual(cell.number_format, "0.00%",
                             "Cell[{}:{}]".format(row, column))

    def test_view_rate_are_percent_formatted(self):
        self.create_test_user()
        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        f = io.BytesIO(response.content)
        book = load_workbook(f)
        sheet = book.worksheets[0]

        row_indexes = range(14, 17)
        column_indexes = [4]
        cell_indexes = product(row_indexes, column_indexes)
        for row, column in cell_indexes:
            cell = sheet[row][column]
            self.assertEqual(cell.number_format, "0.00%",
                             "Cell[{}:{}]".format(row, column))

    def test_ctr_are_percent_formatted(self):
        self.create_test_user()
        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        f = io.BytesIO(response.content)
        book = load_workbook(f)
        sheet = book.worksheets[0]

        row_indexes = range(14, 17)
        column_indexes = [6]
        cell_indexes = product(row_indexes, column_indexes)
        for row, column in cell_indexes:
            cell = sheet[row][column]
            self.assertEqual(cell.number_format, "0.00%",
                             "Cell[{}:{}]".format(row, column))

    def test_demo_data(self):
        user = self.create_test_user()
        account = Account.objects.create(id=next(int_iterator))
        account_creation = AccountCreation.objects.create(owner=user,
                                                          is_managed=False,
                                                          account=account)
        campaign_name = "Test campaign"
        Campaign.objects.create(name=campaign_name)

        url = self._get_url(account_creation.id)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        is_demo_report(sheet)


class AnalyzeExportDashboardAPITestCase(AnalyzeExportAPITestCase):
    def _request(self, account_creation_id):
        url = self._get_url(account_creation_id)
        dashboard_payload = {
            DashboardRequest.DASHBOARD_PARAM_NAME: DashboardRequest.DASHBOARD_PARAM_VALUE,
        }
        return self.client.post(url, json.dumps(dashboard_payload), content_type="application/json")

    def test_no_demo_data(self):
        user = self.create_test_user()
        account = Account.objects.create(id=next(int_iterator))
        account_creation = AccountCreation.objects.create(owner=user,
                                                          is_managed=False,
                                                          account=account)
        campaign_name = "Test campaign"
        Campaign.objects.create(name=campaign_name)

        response = self._request(account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertTrue(is_report_empty(sheet))


def get_sheet_from_response(response):
    single_sheet_index = 0
    f = io.BytesIO(response.content)
    book = load_workbook(f)
    return book.worksheets[single_sheet_index]


TITLE_COLUMN = 1
TOTAL_NAME = "Total"
FIRST_SECTION_ROW_NUMBER = 13


def get_title_cell(sheet):
    return sheet[5][1]


def is_title_empty(sheet):
    title_cell = get_title_cell(sheet)
    return re.match(r"^Campaign: (None)?\n.*", title_cell.value) is not None


def is_demo_report(sheet):
    title_cell = get_title_cell(sheet)
    return re.match(r"^Campaign: Demo\n.*", title_cell.value) is not None and not is_title_empty(sheet)


def is_report_empty(sheet):
    return is_title_empty(sheet) and are_all_sections_empty(sheet)


def are_all_sections_empty(sheet):
    section_names = (
        ("Placement", "Total"),
        ("Ad Groups", None),
        ("Interests", None),
        ("Topics", None),

        ("Keywords", None),
        ("Device", FOOTER_ANNOTATION),
    )
    return all([
        is_section_empty(sheet, section_name, next_value)
        for section_name, next_value in section_names
    ])


def is_section_empty(sheet, section_name, next_value):
    section_row_number = get_section_start_row(sheet, section_name)
    next_row_number = section_row_number + 1
    return sheet[next_row_number][TITLE_COLUMN].value == next_value


def get_section_start_row(sheet, section_name):
    data_rows_numbers = range(FIRST_SECTION_ROW_NUMBER, sheet.max_row)
    for row_number in data_rows_numbers:
        if sheet[row_number][TITLE_COLUMN].value == section_name:
            return row_number
    raise ValueError

