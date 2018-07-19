import io
import json
from datetime import datetime, timedelta, date
from itertools import product, chain
from unittest.mock import patch

from django.core.urlresolvers import reverse
from openpyxl import load_workbook
from rest_framework.status import HTTP_200_OK

from aw_creation.api.urls.names import Name
from aw_creation.models import AccountCreation
from aw_reporting.api.constants import DashboardRequest
from aw_reporting.calculations.cost import get_client_cost
from aw_reporting.demo.models import DEMO_ACCOUNT_ID
from aw_reporting.models import Account, Campaign, AdGroup, AdGroupStatistic, \
    GenderStatistic, AgeRangeStatistic, \
    AudienceStatistic, VideoCreativeStatistic, YTVideoStatistic, \
    YTChannelStatistic, TopicStatistic, \
    KeywordStatistic, CityStatistic, AdStatistic, VideoCreative, GeoTarget, \
    Audience, Topic, Ad, \
    AWConnectionToUserRelation, AWConnection, Opportunity, OpPlacement, SalesForceGoalType
from saas.urls.namespaces import Namespace
from userprofile.models import UserSettingsKey
from utils.utils_tests import ExtendedAPITestCase, int_iterator
from utils.utils_tests import SingleDatabaseApiConnectorPatcher


class PerformanceExportAPITestCase(ExtendedAPITestCase):
    def _get_url(self, account_creation_id):
        return reverse(
            Namespace.AW_CREATION + ":" + Name.Dashboard.PERFORMANCE_EXPORT,
            args=(account_creation_id,))

    def _request(self, account_creation_id, **kwargs):
        url = self._get_url(account_creation_id)
        return self.client.post(url, json.dumps(kwargs), content_type="application/json", )

    def _hide_demo_data_fallback(self, user):
        AWConnectionToUserRelation.objects.create(
            # user must have a connected account not to see demo data
            connection=AWConnection.objects.create(email="me@mail.kz",
                                                   refresh_token=""),
            user=user,
        )

    @staticmethod
    def create_stats(account):
        campaign1 = Campaign.objects.create(id=1, name="#1", account=account)
        ad_group1 = AdGroup.objects.create(id=1, name="", campaign=campaign1)
        campaign2 = Campaign.objects.create(id=2, name="#2", account=account)
        ad_group2 = AdGroup.objects.create(id=2, name="", campaign=campaign2)
        date = datetime.now().date() - timedelta(days=1)
        base_stats = dict(date=date, impressions=100, video_views=10, cost=1)
        topic, _ = Topic.objects.get_or_create(id=1, defaults=dict(name="boo"))
        audience, _ = Audience.objects.get_or_create(id=1,
                                                     defaults=dict(name="boo",
                                                                   type="A"))
        creative, _ = VideoCreative.objects.get_or_create(id=1)
        city, _ = GeoTarget.objects.get_or_create(id=1, defaults=dict(
            name="bobruisk"))
        ad = Ad.objects.create(id=1, ad_group=ad_group1)
        AdStatistic.objects.create(ad=ad, average_position=1, **base_stats)

        for ad_group in (ad_group1, ad_group2):
            stats = dict(ad_group=ad_group, **base_stats)
            AdGroupStatistic.objects.create(average_position=1, **stats)
            GenderStatistic.objects.create(**stats)
            AgeRangeStatistic.objects.create(**stats)
            TopicStatistic.objects.create(topic=topic, **stats)
            AudienceStatistic.objects.create(audience=audience, **stats)
            VideoCreativeStatistic.objects.create(creative=creative, **stats)
            YTChannelStatistic.objects.create(yt_id="123", **stats)
            YTVideoStatistic.objects.create(yt_id="123", **stats)
            KeywordStatistic.objects.create(keyword="123", **stats)
            CityStatistic.objects.create(city=city, **stats)


class PerformanceExportAnalyticsAPITestCase(PerformanceExportAPITestCase):
    def _request(self, account_creation_id, **kwargs):
        self.assertNotIn(DashboardRequest.DASHBOARD_PARAM_NAME, kwargs.keys(),
                         "This test case is for Analytics only. Move this test to appropriate test case: {}"
                         .format(PerformanceExportDashboardAPITestCase.__name__))
        return super(PerformanceExportAnalyticsAPITestCase, self)._request(account_creation_id, **kwargs)

    def assert_demo_data(self, response):
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)

        self.assertFalse(is_empty_report(sheet))

        self.assertEqual(sheet[2][0].value, "Summary")
        for column in range(11, 11 + 4):
            self.assertIsNotNone(sheet[2][column].value, column)

        self.assertIsNotNone(sheet[3][0].value)

    def test_success(self):
        user = self.create_test_user()
        self._hide_demo_data_fallback(user)
        account = Account.objects.create(id=1, name="")
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account,
                                                          is_approved=True)
        self.create_stats(account)

        today = datetime.now().date()
        filters = dict(
            start_date=str(today - timedelta(days=1)),
            end_date=str(today)
        )
        with patch("aw_reporting.charts.SingleDatabaseApiConnector",
                   new=SingleDatabaseApiConnectorPatcher):
            response = self._request(account_creation.id, **filters)
            sheet = get_sheet_from_response(response)
            self.assertFalse(is_empty_report(sheet))

    def test_success_demo(self):
        self.create_test_user()

        today = datetime.now().date()
        filters = dict(
            start_date=str(today - timedelta(days=1)),
            end_date=str(today)
        )
        response = self._request(DEMO_ACCOUNT_ID, **filters)
        self.assert_demo_data(response)

    def test_report_is_xlsx_formatted(self):
        user = self.create_test_user()
        self._hide_demo_data_fallback(user)
        account = Account.objects.create(id=1, name="")
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account,
                                                          is_approved=True)
        self.create_stats(account)

        with patch("aw_reporting.charts.SingleDatabaseApiConnector",
                   new=SingleDatabaseApiConnectorPatcher):
            response = self._request(account_creation.id)
            self.assertEqual(response.status_code, HTTP_200_OK)
            try:
                f = io.BytesIO(response.content)
                load_workbook(f)
            except:
                self.fail("Report is not an xls")

    def test_report_percent_formatted(self):
        user = self.create_test_user()
        self._hide_demo_data_fallback(user)
        account = Account.objects.create(id=1, name="")
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account,
                                                          is_approved=True)
        self.create_stats(account)

        with patch("aw_reporting.charts.SingleDatabaseApiConnector",
                   new=SingleDatabaseApiConnectorPatcher):
            response = self._request(account_creation.id)
            self.assertEqual(response.status_code, HTTP_200_OK)
            sheet = get_sheet_from_response(response)
            self.assertFalse(is_empty_report(sheet))
            rows = range(2, sheet.max_row + 1)
            ctr_range = 8, 10,
            view_rate_range = 10, 11
            quartile_range = 11, 15
            test_ranges = [range(start, end) for start, end
                           in [ctr_range, view_rate_range, quartile_range]]
            cols = chain(*test_ranges)
            test_indexes = product(rows, cols)
            for row, column in test_indexes:
                cell = sheet[row][column]
                self.assertEqual(cell.number_format, "0.00%",
                                 "Cell[{}:{}]".format(row, column))

    def test_demo_data_fallback(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=1, name="")
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account,
                                                          is_approved=True)

        campaign_name = "Test campaign"
        Campaign.objects.create(name=campaign_name)

        response = self._request(account_creation.id)
        self.assert_demo_data(response)


class PerformanceExportDashboardAPITestCase(PerformanceExportAPITestCase):
    def _request(self, account_creation_id, **kwargs):
        kwargs[DashboardRequest.DASHBOARD_PARAM_NAME] = DashboardRequest.DASHBOARD_PARAM_VALUE
        return super(PerformanceExportDashboardAPITestCase, self)._request(account_creation_id, **kwargs)

    def test_success_for_chf_dashboard(self):
        user = self.create_test_user()
        user.is_staff = True
        user.save()
        self._hide_demo_data_fallback(user)
        account = Account.objects.create(id=1, name="")
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account,
                                                          is_approved=True)
        self.create_stats(account)
        user_settings = {
            UserSettingsKey.VISIBLE_ACCOUNTS: [1],
        }
        with patch("aw_reporting.charts.SingleDatabaseApiConnector",
                   new=SingleDatabaseApiConnectorPatcher), \
             self.patch_user_settings(**user_settings):
            response = self._request(account_creation.id)
            self.assertEqual(response.status_code, HTTP_200_OK)

    def test_no_demo_data(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=1, name="")
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account,
                                                          is_approved=True)

        campaign_name = "Test campaign"
        Campaign.objects.create(name=campaign_name)
        user_settings = {
            UserSettingsKey.VISIBLE_ACCOUNTS: [1],
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertTrue(is_empty_report(sheet))

    def test_sf_data_in_summary_row(self):
        user = self.create_test_user()
        any_date = date(2018, 1, 1)
        user.add_custom_user_permission("view_dashboard")
        opportunity = Opportunity.objects.create()
        placement = OpPlacement.objects.create(opportunity=opportunity, ordered_rate=.2, total_cost=23,
                                               goal_type_id=SalesForceGoalType.CPM)

        account = Account.objects.create(id=next(int_iterator), name="")
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account,
                                                          is_approved=True)
        campaign = Campaign.objects.create(name="", account=account, salesforce_placement=placement)
        ad_group = AdGroup.objects.create(campaign=campaign)
        impressions, views, aw_cost = 1234, 234, 12
        AdGroupStatistic.objects.create(date=any_date, ad_group=ad_group, average_position=1,
                                        cost=aw_cost, impressions=impressions, video_views=views)
        client_cost = get_client_cost(
            goal_type_id=placement.goal_type_id,
            dynamic_placement=placement.dynamic_placement,
            placement_type=placement.placement_type,
            ordered_rate=placement.ordered_rate,
            impressions=impressions,
            video_views=views,
            aw_cost=aw_cost,
            total_cost=placement.total_cost,
            tech_fee=placement.tech_fee,
            start=any_date,
            end=any_date
        )
        self.assertGreater(client_cost, 0)
        average_cpm = client_cost / impressions * 1000
        average_cpv = client_cost / views
        user_settings = {
            UserSettingsKey.VISIBLE_ACCOUNTS: [account.id],
            UserSettingsKey.DASHBOARD_AD_WORDS_RATES: False
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))

        self.assertAlmostEqual(sheet[SUMMARY_ROW_NUMBER][Column.COST].value, client_cost)
        self.assertAlmostEqual(sheet[SUMMARY_ROW_NUMBER][Column.AVERAGE_CPM].value, average_cpm)
        self.assertAlmostEqual(sheet[SUMMARY_ROW_NUMBER][Column.AVERAGE_CPV].value, average_cpv)


def get_sheet_from_response(response):
    single_sheet_index = 0
    f = io.BytesIO(response.content)
    book = load_workbook(f)
    return book.worksheets[single_sheet_index]


SUMMARY_ROW_NUMBER = 2


class Column:
    IMPRESSIONS = 2
    VIEWS = 3
    COST = 4
    AVERAGE_CPM = 5
    AVERAGE_CPV = 6
    CLICKS = 7
    CTR_I = 8
    CTR_V = 9
    VIEW_RATE = 10
    QUARTERS = range(11, 15)


def is_summary_empty(sheet):
    values_columns_indexes = range(2, 15)
    return all([sheet[SUMMARY_ROW_NUMBER][column].value is None for column in values_columns_indexes])


def is_empty_report(sheet):
    min_rows_count = 2
    return sheet.max_row <= min_rows_count and is_summary_empty(sheet)
