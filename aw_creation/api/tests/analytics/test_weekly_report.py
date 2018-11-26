import io
import re
from functools import partial
from itertools import product

from openpyxl import load_workbook
from rest_framework.status import HTTP_200_OK

from aw_creation.api.urls.names import Name
from aw_creation.api.urls.namespace import Namespace
from aw_creation.models import AccountCreation
from aw_reporting.demo.models import DEMO_ACCOUNT_ID, DemoAccount
from aw_reporting.models import Account
from aw_reporting.models import Campaign
from saas.urls.namespaces import Namespace as RootNamespace
from utils.utittests.generic_test import generic_test
from utils.utittests.int_iterator import int_iterator
from utils.utittests.reverse import reverse
from utils.utittests.test_case import ExtendedAPITestCase


class SectionName:
    AD_GROUPS = "Ad Groups"
    PLACEMENT = "Placement"
    INTERESTS = "Interests"
    TOPICS = "Topics"
    DEVICES = "Device"
    KEYWORDS = "Keywords"


STATISTIC_SECTIONS = (
    SectionName.AD_GROUPS,
    SectionName.DEVICES,
    SectionName.INTERESTS,
    SectionName.KEYWORDS,
    SectionName.PLACEMENT,
    SectionName.TOPICS,
)


class Column:
    CLICKS = "Clicks"
    CTA_APP_STORE = "App Store"
    CTA_CARDS = "Cards"
    CTA_END_CAP = "End cap"
    CTA_OVERLAY = "Call-to-Action overlay"
    CTA_WEBSITE = "Website"
    CTR = "CTR"
    IMPRESSIONS = "Impressions"
    VIDEO_PLAYED_PERCENT_100 = "Video played to: 100%"
    VIDEO_PLAYED_PERCENT_25 = "Video played to: 25%"
    VIDEO_PLAYED_PERCENT_50 = "Video played to: 50%"
    VIDEO_PLAYED_PERCENT_75 = "Video played to: 75%"
    VIEW_RATE = "View Rate"
    VIEWS = "Views"


COLUMNS_ORDER = (
    Column.IMPRESSIONS,
    Column.VIEWS,
    Column.VIEW_RATE,
    Column.CLICKS,
    Column.CTA_OVERLAY,
    Column.CTA_WEBSITE,
    Column.CTA_APP_STORE,
    Column.CTA_CARDS,
    Column.CTA_END_CAP,
    Column.CTR,
    Column.VIDEO_PLAYED_PERCENT_25,
    Column.VIDEO_PLAYED_PERCENT_50,
    Column.VIDEO_PLAYED_PERCENT_75,
    Column.VIDEO_PLAYED_PERCENT_100,
)


class AnalyticsWeeklyReportAPITestCase(ExtendedAPITestCase):
    def _get_url(self, account_creation_id):
        return reverse(
            Name.Analytics.PERFORMANCE_EXPORT_WEEKLY_REPORT,
            [RootNamespace.AW_CREATION, Namespace.ANALYTICS],
            args=(account_creation_id,)
        )

    def test_success(self):
        user = self.create_test_user()
        account = Account.objects.create(id=1, name="",
                                         skip_creating_account_creation=True)
        account_creation = AccountCreation.objects.create(name="", owner=user,
                                                          is_managed=False,
                                                          account=account)

        url = self._get_url(account_creation.id)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)

    def test_success_demo(self):
        self.create_test_user()
        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)

    def test_quartiles_are_percent_formatted(self):
        self.create_test_user()
        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        f = io.BytesIO(response.content)
        book = load_workbook(f)
        sheet = book.worksheets[0]

        row_indexes = range(14, 17)
        column_indexes = range(7, 11)
        cell_indexes = product(row_indexes, column_indexes)
        for row, column in cell_indexes:
            cell = sheet[row][column]
            self.assertEqual(cell.number_format, "0.00%",
                             "Cell[{}:{}]".format(row, column))

    def test_view_rate_are_percent_formatted(self):
        self.create_test_user()
        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        f = io.BytesIO(response.content)
        book = load_workbook(f)
        sheet = book.worksheets[0]

        row_indexes = range(14, 17)
        column_indexes = [4]
        cell_indexes = product(row_indexes, column_indexes)
        for row, column in cell_indexes:
            cell = sheet[row][column]
            self.assertEqual(cell.number_format, "0.00%",
                             "Cell[{}:{}]".format(row, column))

    def test_ctr_are_percent_formatted(self):
        self.create_test_user()
        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        f = io.BytesIO(response.content)
        book = load_workbook(f)
        sheet = book.worksheets[0]

        row_indexes = range(14, 17)
        column_indexes = [6]
        cell_indexes = product(row_indexes, column_indexes)
        for row, column in cell_indexes:
            cell = sheet[row][column]
            self.assertEqual(cell.number_format, "0.00%",
                             "Cell[{}:{}]".format(row, column))

    def test_demo_data(self):
        user = self.create_test_user()
        account = Account.objects.create(id=next(int_iterator),
                                         skip_creating_account_creation=True)
        account_creation = AccountCreation.objects.create(owner=user,
                                                          is_managed=False,
                                                          account=account)
        campaign_name = "Test campaign"
        Campaign.objects.create(name=campaign_name)

        url = self._get_url(account_creation.id)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        is_demo_report(sheet)

    @generic_test([
        (section, (section,), dict())
        for section in STATISTIC_SECTIONS
    ])
    def test_column_set(self, section):
        user = self.create_test_user()
        account = Account.objects.create(id=next(int_iterator),
                                         skip_creating_account_creation=True)
        account_creation = AccountCreation.objects.create(owner=user,
                                                          is_managed=False,
                                                          account=account)

        url = self._get_url(account_creation.id)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        row_index = get_section_start_row(sheet, section)
        title_values = tuple(cell.value for cell in sheet[row_index][1:])
        self.assertEqual(title_values, (section,) + COLUMNS_ORDER)

    @generic_test([
        (section, (section,), dict())
        for section in STATISTIC_SECTIONS
    ])
    def test_demo_account_cta(self, section):
        self.create_test_user()

        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        row_index = get_section_start_row(sheet, section)
        title_values = tuple(cell.value for cell in sheet[row_index][1:])
        self.assertEqual(title_values, (section,) + COLUMNS_ORDER)

    def test_demo_total_contains_cta(self):
        self.create_test_user()
        demo_account = DemoAccount()

        url = self._get_url(DEMO_ACCOUNT_ID)
        response = self.client.post(url)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        total_row = get_total_row(sheet)
        columns = get_section_columns(sheet, SectionName.PLACEMENT)
        total_value = partial(get_value_in_column, total_row, columns)
        self.assertEqual(total_value(Column.CTA_CARDS), demo_account.clicks_cards)
        self.assertEqual(total_value(Column.CTA_APP_STORE), demo_account.clicks_app_store)
        self.assertEqual(total_value(Column.CTA_WEBSITE), demo_account.clicks_website)
        self.assertEqual(total_value(Column.CTA_END_CAP), demo_account.clicks_end_cap)
        self.assertEqual(total_value(Column.CTA_OVERLAY), demo_account.clicks_call_to_action_overlay)


TITLE_COLUMN = 1
FIRST_SECTION_ROW_NUMBER = 13


def get_sheet_from_response(response):
    single_sheet_index = 0
    f = io.BytesIO(response.content)
    book = load_workbook(f)
    return book.worksheets[single_sheet_index]


def get_title_cell(sheet):
    return sheet[5][1]


def is_title_empty(sheet):
    title_cell = get_title_cell(sheet)
    return re.match(r"^Campaign: (None)?\n.*", title_cell.value) is not None


def is_demo_report(sheet):
    title_cell = get_title_cell(sheet)
    return re.match(r"^Campaign: Demo\n.*", title_cell.value) is not None and not is_title_empty(sheet)


def is_section_header(cell):
    return cell.alignment.horizontal == "center"


def get_section_start_row(sheet, section_name):
    data_rows_numbers = range(FIRST_SECTION_ROW_NUMBER, sheet.max_row)
    for row_number in data_rows_numbers:
        cell = sheet[row_number][TITLE_COLUMN]
        if is_section_header(cell) and cell.value == section_name:
            return row_number
    raise ValueError


def get_total_row(sheet):
    placement_title_row = get_section_start_row(sheet, SectionName.PLACEMENT)
    placement_total_row = placement_title_row + 1
    while sheet[placement_total_row][TITLE_COLUMN].value != "Total":
        placement_total_row += 1
    return sheet[placement_total_row]


def get_section_columns(sheet, section_name):
    section_index = get_section_start_row(sheet, section_name)
    return [cell.value for cell in sheet[section_index][TITLE_COLUMN:]]


def get_value_in_column(row, headers, column_name):
    column_index = headers.index(column_name)
    return row[TITLE_COLUMN + column_index].value
