import io
import json
from datetime import date
from datetime import datetime
from datetime import timedelta
from unittest.mock import patch

from django.test import override_settings
from openpyxl import load_workbook
from rest_framework.status import HTTP_200_OK
from rest_framework.status import HTTP_400_BAD_REQUEST

from aw_creation.api.urls.names import Name
from aw_creation.api.urls.namespace import Namespace
from aw_creation.api.views.dashboard.performance_export import METRIC_REPRESENTATION
from aw_creation.api.views.dashboard.performance_export import Metric
from aw_reporting.calculations.cost import get_client_cost
from aw_reporting.dashboard_charts import DateSegment
from aw_reporting.excel_reports.dashboard_performance_report import COLUMN_NAME, TOO_MUCH_DATA_MESSAGE
from aw_reporting.excel_reports.dashboard_performance_report import DashboardPerformanceReportColumn
from aw_reporting.models import Account
from aw_reporting.models import Ad
from aw_reporting.models import AdGroup
from aw_reporting.models import AdGroupStatistic
from aw_reporting.models import AdStatistic
from aw_reporting.models import AgeRangeStatistic
from aw_reporting.models import Audience
from aw_reporting.models import AudienceStatistic
from aw_reporting.models import Campaign
from aw_reporting.models import CityStatistic
from aw_reporting.models import GenderStatistic
from aw_reporting.models import GeoTarget
from aw_reporting.models import KeywordStatistic
from aw_reporting.models import OpPlacement
from aw_reporting.models import Opportunity
from aw_reporting.models import RemarkList
from aw_reporting.models import RemarkStatistic
from aw_reporting.models import SalesForceGoalType
from aw_reporting.models import Topic
from aw_reporting.models import TopicStatistic
from aw_reporting.models import VideoCreative
from aw_reporting.models import VideoCreativeStatistic
from aw_reporting.models import YTChannelStatistic
from aw_reporting.models import YTVideoStatistic
from saas.urls.namespaces import Namespace as RootNamespace
from userprofile.constants import UserSettingsKey
from utils.utils_tests import ExtendedAPITestCase
from utils.utils_tests import SingleDatabaseApiConnectorPatcher
from utils.utils_tests import generic_test
from utils.utils_tests import int_iterator
from utils.utils_tests import reverse


class DashboardPerformanceExportAPITestCase(ExtendedAPITestCase):
    def _get_url(self, account_creation_id):
        return reverse(Name.Dashboard.PERFORMANCE_EXPORT, [RootNamespace.AW_CREATION, Namespace.DASHBOARD],
                       args=(account_creation_id,))

    def _request(self, account_creation_id, **kwargs):
        url = self._get_url(account_creation_id)
        return self.client.post(url, json.dumps(kwargs), content_type="application/json", )

    def _create_stats(self, account, statistic_date=None):
        campaign1 = Campaign.objects.create(id=1, name="#1", account=account)
        ad_group1 = AdGroup.objects.create(id=1, name="", campaign=campaign1)
        campaign2 = Campaign.objects.create(id=2, name="#2", account=account)
        ad_group2 = AdGroup.objects.create(id=2, name="", campaign=campaign2)
        statistic_date = statistic_date or (datetime.now().date() - timedelta(days=1))
        base_stats = dict(date=statistic_date, impressions=100, video_views=10, cost=1)
        topic, _ = Topic.objects.get_or_create(id=1, defaults=dict(name="boo"))
        audience, _ = Audience.objects.get_or_create(id=1,
                                                     defaults=dict(name="boo",
                                                                   type="A"))
        remark, _ = RemarkList.objects.get_or_create(name="test")
        creative, _ = VideoCreative.objects.get_or_create(id=1)
        city, _ = GeoTarget.objects.get_or_create(id=1, defaults=dict(
            name="bobruisk"))
        ad = Ad.objects.create(id=1, ad_group=ad_group1)
        AdStatistic.objects.create(ad=ad, average_position=1, **base_stats)

        for ad_group in (ad_group1, ad_group2):
            stats = dict(ad_group=ad_group, **base_stats)
            AdGroupStatistic.objects.create(average_position=1, **stats)
            GenderStatistic.objects.create(**stats)
            AgeRangeStatistic.objects.create(**stats)
            TopicStatistic.objects.create(topic=topic, **stats)
            AudienceStatistic.objects.create(audience=audience, **stats)
            VideoCreativeStatistic.objects.create(creative=creative, **stats)
            YTChannelStatistic.objects.create(yt_id="123", **stats)
            YTVideoStatistic.objects.create(yt_id="123", **stats)
            KeywordStatistic.objects.create(keyword="123", **stats)
            CityStatistic.objects.create(city=city, **stats)
            RemarkStatistic.objects.create(remark=remark, **stats)

    def test_success_for_chf_dashboard(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=1, name="")
        self._create_stats(account)
        user_settings = {
            UserSettingsKey.VISIBLE_ACCOUNTS: [1],
        }
        with patch("aw_reporting.dashboard_charts.SingleDatabaseApiConnector",
                   new=SingleDatabaseApiConnectorPatcher), \
             self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
            self.assertEqual(response.status_code, HTTP_200_OK)

    def test_no_demo_data(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=1, name="")

        campaign_name = "Test campaign"
        Campaign.objects.create(name=campaign_name)
        user_settings = {
            UserSettingsKey.VISIBLE_ACCOUNTS: [1],
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertTrue(is_empty_report(sheet))

    def test_sf_data_in_summary_row(self):
        user = self.create_test_user()
        any_date = date(2018, 1, 1)
        user.add_custom_user_permission("view_dashboard")
        opportunity = Opportunity.objects.create()
        placement = OpPlacement.objects.create(opportunity=opportunity, ordered_rate=.2, total_cost=23,
                                               goal_type_id=SalesForceGoalType.CPM)

        account = Account.objects.create(id=next(int_iterator), name="")
        campaign = Campaign.objects.create(name="", account=account, salesforce_placement=placement)
        ad_group = AdGroup.objects.create(campaign=campaign)
        impressions, views, aw_cost = 1234, 234, 12
        AdGroupStatistic.objects.create(date=any_date, ad_group=ad_group, average_position=1,
                                        cost=aw_cost, impressions=impressions, video_views=views)
        client_cost = get_client_cost(
            goal_type_id=placement.goal_type_id,
            dynamic_placement=placement.dynamic_placement,
            placement_type=placement.placement_type,
            ordered_rate=placement.ordered_rate,
            impressions=impressions,
            video_views=views,
            aw_cost=aw_cost,
            total_cost=placement.total_cost,
            tech_fee=placement.tech_fee,
            start=any_date,
            end=any_date
        )
        self.assertGreater(client_cost, 0)
        average_cpm = client_cost / impressions * 1000
        average_cpv = client_cost / views
        user_settings = {
            UserSettingsKey.VISIBLE_ACCOUNTS: [account.id],
            UserSettingsKey.DASHBOARD_AD_WORDS_RATES: False
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))
        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])
        cost_index = get_column_index(headers, DashboardPerformanceReportColumn.COST)
        cpm_index = get_column_index(headers, DashboardPerformanceReportColumn.AVERAGE_CPM)
        cpv_index = get_column_index(headers, DashboardPerformanceReportColumn.AVERAGE_CPV)
        self.assertAlmostEqual(sheet[SUMMARY_ROW_INDEX][cost_index].value, client_cost)
        self.assertAlmostEqual(sheet[SUMMARY_ROW_INDEX + 1][cost_index].value, client_cost)
        self.assertAlmostEqual(sheet[SUMMARY_ROW_INDEX][cpm_index].value, average_cpm)
        self.assertAlmostEqual(sheet[SUMMARY_ROW_INDEX][cpv_index].value, average_cpv)

    def test_hide_costs(self):
        user = self.create_test_user()
        any_date = date(2018, 1, 1)
        user.add_custom_user_permission("view_dashboard")
        opportunity = Opportunity.objects.create()
        placement = OpPlacement.objects.create(opportunity=opportunity, ordered_rate=.2, total_cost=23,
                                               goal_type_id=SalesForceGoalType.CPM)

        account = Account.objects.create(id=next(int_iterator), name="")
        campaign = Campaign.objects.create(name="", account=account, salesforce_placement=placement)
        ad_group = AdGroup.objects.create(campaign=campaign)
        AdGroupStatistic.objects.create(date=any_date, ad_group=ad_group, average_position=1,
                                        cost=1, impressions=1, video_views=1)
        user_settings = {
            UserSettingsKey.VISIBLE_ACCOUNTS: [account.id],
            UserSettingsKey.DASHBOARD_COSTS_ARE_HIDDEN: True
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))
        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])
        expected_headers = (
            None, "Name", "Impressions", "Views", "Clicks", "Call-to-Action overlay", "Website", "App Store", "Cards",
            "End cap", "Ctr(i)", "Ctr(v)", "View rate", "25%", "50%", "75%", "100%")

        self.assertEqual(headers, expected_headers)
        row_lengths = [len(row) for row in sheet.rows]
        self.assertTrue(all([length == len(expected_headers) for length in row_lengths]))

    def test_all_conversions_column_is_present(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        any_date = date(2018, 1, 1)
        opportunity = Opportunity.objects.create()
        placement = OpPlacement.objects.create(
            opportunity=opportunity, ordered_rate=.2, total_cost=23, goal_type_id=SalesForceGoalType.CPM)
        account = Account.objects.create(id=next(int_iterator), name="")
        campaign = Campaign.objects.create(name="", account=account, salesforce_placement=placement)
        ad_group = AdGroup.objects.create(campaign=campaign)
        AdGroupStatistic.objects.create(
            date=any_date, ad_group=ad_group, average_position=1, cost=1, impressions=1, video_views=1)
        user_settings = {
            UserSettingsKey.VISIBLE_ACCOUNTS: [account.id],
            UserSettingsKey.SHOW_CONVERSIONS: True
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))
        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])
        expected_headers = (
            None, "Name", "Impressions", "Views", "Cost", "Average cpm", "Average cpv", "Clicks",
            "Call-to-Action overlay", "Website", "App Store", "Cards", "End cap", "Ctr(i)", "Ctr(v)", "View rate",
            "25%", "50%", "75%", "100%", "All conversions")
        self.assertEqual(headers, expected_headers)
        row_lengths = [len(row) for row in sheet.rows]
        self.assertTrue(all([length == len(expected_headers) for length in row_lengths]))

    def test_show_real_costs(self):
        user = self.create_test_user()
        any_date = date(2018, 1, 1)
        user.add_custom_user_permission("view_dashboard")
        opportunity = Opportunity.objects.create()
        placement = OpPlacement.objects.create(opportunity=opportunity, ordered_rate=.2, total_cost=23,
                                               goal_type_id=SalesForceGoalType.CPM)

        account = Account.objects.create(id=next(int_iterator), name="")
        campaign = Campaign.objects.create(name="", account=account, salesforce_placement=placement)
        ad_group = AdGroup.objects.create(campaign=campaign)
        aw_cost = 123
        AdGroupStatistic.objects.create(date=any_date, ad_group=ad_group, average_position=1,
                                        cost=aw_cost, impressions=1, video_views=1)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
            UserSettingsKey.DASHBOARD_AD_WORDS_RATES: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))
        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])
        cost_column_index = headers.index("Cost")
        single_data_row_index = 3
        self.assertEqual(sheet[single_data_row_index][cost_column_index].value, aw_cost)

    def test_hide_real_costs(self):
        user = self.create_test_user()
        any_date = date(2018, 1, 1)
        user.add_custom_user_permission("view_dashboard")
        opportunity = Opportunity.objects.create()
        placement = OpPlacement.objects.create(opportunity=opportunity, ordered_rate=.2, total_cost=23,
                                               goal_type_id=SalesForceGoalType.CPM)

        account = Account.objects.create(id=next(int_iterator), name="")
        campaign = Campaign.objects.create(name="", account=account, salesforce_placement=placement)
        ad_group = AdGroup.objects.create(campaign=campaign)
        stats = AdGroupStatistic.objects.create(date=any_date, ad_group=ad_group, average_position=1,
                                                cost=123, impressions=1, video_views=1)
        client_cost = get_client_cost(
            goal_type_id=placement.goal_type_id,
            dynamic_placement=placement.dynamic_placement,
            placement_type=placement.placement_type,
            ordered_rate=placement.ordered_rate,
            impressions=stats.impressions,
            video_views=stats.video_views,
            aw_cost=stats.cost,
            total_cost=placement.total_cost,
            tech_fee=placement.tech_fee,
            start=any_date,
            end=any_date,
        )
        self.assertGreater(client_cost, 0)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
            UserSettingsKey.DASHBOARD_AD_WORDS_RATES: False,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))
        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])
        cost_column_index = headers.index("Cost")
        single_data_row_index = 3
        self.assertEqual(sheet[single_data_row_index][cost_column_index].value, client_cost)

    @generic_test([
        (metric, (metric,), dict())
        for metric in Metric
    ])
    def test_metric(self, metric):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")

        account = Account.objects.create(id=next(int_iterator), name="")
        self._create_stats(account)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
            UserSettingsKey.DASHBOARD_COSTS_ARE_HIDDEN: True
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, metric=metric.value)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))
        data_rows = list(sheet.rows)[SUMMARY_ROW_INDEX:]
        self.assertEqual(len(data_rows), 1)
        self.assertEqual(data_rows[0][0].value, METRIC_REPRESENTATION[metric])

    def test_bad_request_on_wrong_metric(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")

        account = Account.objects.create(id=next(int_iterator), name="")
        self._create_stats(account)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
            UserSettingsKey.DASHBOARD_COSTS_ARE_HIDDEN: True
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, metric="overview")
        self.assertEqual(response.status_code, HTTP_400_BAD_REQUEST)

    def test_date_segment_invalid(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        test_segment = "invalid"
        self.assertFalse(DateSegment.has_value(test_segment))
        account = Account.objects.create(id=next(int_iterator), name="")
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, date_segment=test_segment)
        self.assertEqual(response.status_code, HTTP_400_BAD_REQUEST)

    def test_date_segment_day(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        test_date = datetime(2018, 4, 5, 6, 7, 8, 9)
        self._create_stats(account, test_date)
        expected_date_label = test_date.strftime("%m/%d/%Y")
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, date_segment=DateSegment.DAY.value)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))

        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])

        self.assertEqual(headers[1], "Date")
        data_rows = list(sheet.rows)[SUMMARY_ROW_INDEX:]
        self.assertEqual(data_rows[0][1].value, expected_date_label)

    def test_date_segment_week(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        test_date = datetime(2018, 4, 5, 6, 7, 8, 9)
        self._create_stats(account, test_date)
        expected_date_label = test_date.strftime("%V")
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, date_segment=DateSegment.WEEK.value)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))

        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])

        self.assertEqual(headers[1], "Date")
        data_rows = list(sheet.rows)[SUMMARY_ROW_INDEX:]
        self.assertEqual(data_rows[0][1].value, expected_date_label)

    def test_date_segment_month(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        test_date = datetime(2018, 4, 5, 6, 7, 8, 9)
        self._create_stats(account, test_date)
        expected_date_label = test_date.strftime("%b-%y")
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, date_segment=DateSegment.MONTH.value)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))

        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])

        self.assertEqual(headers[1], "Date")
        data_rows = list(sheet.rows)[SUMMARY_ROW_INDEX:]
        self.assertEqual(data_rows[0][1].value, expected_date_label)

    def test_date_segment_year(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        test_date = datetime(2018, 4, 5, 6, 7, 8, 9)
        self._create_stats(account, test_date)
        expected_date_label = test_date.strftime("%Y")
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, date_segment=DateSegment.YEAR.value)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))

        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])

        self.assertEqual(headers[1], "Date")
        data_rows = list(sheet.rows)[SUMMARY_ROW_INDEX:]
        self.assertEqual(data_rows[0][1].value, expected_date_label)

    def test_date_segment_split_by_day(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        campaign = Campaign.objects.create(account=account)
        ad_group = AdGroup.objects.create(campaign=campaign)
        test_date_1 = datetime(2018, 4, 5, 6, 7, 8, 9)
        test_date_2 = test_date_1 + timedelta(days=1)
        impressions = (2, 3)
        common = dict(
            average_position=1,
            ad_group=ad_group,
        )
        AdGroupStatistic.objects.create(date=test_date_1, impressions=impressions[0], **common)
        AdGroupStatistic.objects.create(date=test_date_2, impressions=impressions[1], **common)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, date_segment=DateSegment.DAY.value)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertFalse(is_empty_report(sheet))
        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])
        impressions_column = get_column_index(headers, DashboardPerformanceReportColumn.IMPRESSIONS)
        data_rows = list(sheet.rows)[SUMMARY_ROW_INDEX:]

        self.assertEqual(data_rows[0][impressions_column].value, impressions[0])
        self.assertEqual(data_rows[1][impressions_column].value, impressions[1])

    def test_date_segment_empty_in_summary(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        self._create_stats(account)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, date_segment=DateSegment.DAY.value)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        headers = tuple(cell.value for cell in sheet[HEADER_ROW_INDEX])
        date_column = get_column_index(headers, DashboardPerformanceReportColumn.DATE_SEGMENT)
        summary_row = sheet[SUMMARY_ROW_INDEX]
        self.assertIsNone(summary_row[date_column].value)

    def test_filters_header_metric(self):
        user = self.create_test_user()
        test_metric = Metric.AD_GROUP
        test_metric_repr = METRIC_REPRESENTATION[test_metric]
        test_metric_value = test_metric.value
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, metric=test_metric_value)
        sheet = get_sheet_from_response(response)
        header = get_custom_header(sheet)
        self.assertIn("Group By: {}".format(test_metric_repr), header)

    def test_filters_header_date(self):
        user = self.create_test_user()
        test_start_date = "2018-01-01"
        test_end_date = "2018-01-02"
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, start_date=test_start_date, end_date=test_end_date)
        sheet = get_sheet_from_response(response)
        header = get_custom_header(sheet)
        self.assertIn("Date: {} - {}".format(test_start_date, test_end_date), header)

    def test_filters_header_campaigns(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        campaigns = [
            Campaign.objects.create(id=next(int_iterator), name="test name 1", account=account),
            Campaign.objects.create(id=next(int_iterator), name="test name 2", account=account),
        ]
        campaign_ids = [campaign.id for campaign in campaigns]
        campaign_names = [campaign.name for campaign in campaigns]
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, campaigns=campaign_ids)
        sheet = get_sheet_from_response(response)
        header = get_custom_header(sheet)
        self.assertIn("Campaigns: {}".format(", ".join(campaign_names)), header)

    def test_filters_header_ad_groups(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        account = Account.objects.create(id=next(int_iterator), name="")
        campaign = Campaign.objects.create(id=next(int_iterator), account=account)
        ad_groups = [
            AdGroup.objects.create(id=next(int_iterator), name="test name 1", campaign=campaign),
            AdGroup.objects.create(id=next(int_iterator), name="test name 2", campaign=campaign),
        ]
        ad_group_ids = [ad_group.id for ad_group in ad_groups]
        ad_group_names = [ad_group.name for ad_group in ad_groups]
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id, ad_groups=ad_group_ids)
        sheet = get_sheet_from_response(response)
        header = get_custom_header(sheet)
        self.assertIn("Ad Groups: {}".format(", ".join(ad_group_names)), header)

    def test_filters_no_filters(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        expected_header = "\n".join([
            "Date: None - None",
            "Group By: None",
            "Campaigns: ",
            "Ad Groups: ",
        ])
        account = Account.objects.create(id=next(int_iterator), name="")
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        sheet = get_sheet_from_response(response)
        header = get_custom_header(sheet)
        self.assertEqual(header, expected_header)

    def test_too_much_data(self):
        user = self.create_test_user()
        user.add_custom_user_permission("view_dashboard")
        test_data_limit = 1
        test_limit = SUMMARY_ROW_INDEX + test_data_limit
        account = Account.objects.create(id=next(int_iterator), name="")
        self._create_stats(account)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True,
        }
        with self.patch_user_settings(**user_settings), \
             override_settings(DASHBOARD_PERFORMANCE_REPORT_LIMIT=test_limit):
            response = self._request(account.account_creation.id)
        sheet = get_sheet_from_response(response)
        data_rows = list(sheet.rows)[SUMMARY_ROW_INDEX:]
        self.assertEqual(len(data_rows), test_data_limit + 1)
        self.assertEqual(data_rows[test_data_limit][0].value, TOO_MUCH_DATA_MESSAGE)


def get_sheet_from_response(response):
    single_sheet_index = 0
    f = io.BytesIO(response.content)
    book = load_workbook(f)
    return book.worksheets[single_sheet_index]


CUSTOM_HEADER_ROW_INDEX = 1
HEADER_ROW_INDEX = 2
SUMMARY_ROW_INDEX = 3


def is_summary_empty(sheet):
    values_columns_indexes = range(2, 15)
    return all([sheet[SUMMARY_ROW_INDEX][column].value is None for column in values_columns_indexes])


def is_empty_report(sheet):
    min_rows_count = SUMMARY_ROW_INDEX
    return sheet.max_row <= min_rows_count and is_summary_empty(sheet)


def get_column_index(headers, column):
    return headers.index(COLUMN_NAME[column])


def get_custom_header(sheet):
    return sheet[CUSTOM_HEADER_ROW_INDEX][2].value
