import io
import re
from datetime import date
from datetime import timedelta
from unittest.mock import patch

from openpyxl import load_workbook
from rest_framework.status import HTTP_200_OK

from aw_creation.api.urls.names import Name
from aw_creation.api.urls.namespace import Namespace
from aw_reporting.demo.models import DEMO_ACCOUNT_ID
from aw_reporting.excel_reports_dashboard import FOOTER_ANNOTATION
from aw_reporting.models import Account
from aw_reporting.models import AdGroup
from aw_reporting.models import AgeRange
from aw_reporting.models import AgeRangeStatistic
from aw_reporting.models import Audience
from aw_reporting.models import AudienceStatistic
from aw_reporting.models import Campaign
from aw_reporting.models import Gender
from aw_reporting.models import GenderStatistic
from aw_reporting.models import KeywordStatistic
from aw_reporting.models import OpPlacement
from aw_reporting.models import Opportunity
from aw_reporting.models import Topic
from aw_reporting.models import TopicStatistic
from aw_reporting.models import VideoCreative
from aw_reporting.models import VideoCreativeStatistic
from aw_reporting.models import YTChannelStatistic
from aw_reporting.models import YTVideoStatistic
from aw_reporting.models import age_range_str
from aw_reporting.models import gender_str
from saas.urls.namespaces import Namespace as RootNamespace
from userprofile.constants import UserSettingsKey
from utils.utils_tests import ExtendedAPITestCase
from utils.utils_tests import SingleDatabaseApiConnectorPatcher
from utils.utils_tests import generic_test
from utils.utils_tests import int_iterator
from utils.utils_tests import patch_now
from utils.utils_tests import reverse


class SectionName:
    AD_GROUPS = "Ad Groups"
    AGES = "Ages"
    CREATIVES = "Creatives"
    DEVICES = "Device"
    GENDERS = "Genders"
    INTERESTS = "Interests"
    KEYWORDS = "Keywords"
    PLACEMENT = "Placement"
    TOPICS = "Topics"
    VIDEOS = "Video"
    TARGETING_TACTICS = "Targeting Tactic"


SECTIONS_WITH_CTA = (
    SectionName.AD_GROUPS,
    SectionName.AGES,
    SectionName.DEVICES,
    SectionName.GENDERS,
    SectionName.INTERESTS,
    SectionName.KEYWORDS,
    SectionName.PLACEMENT,
    SectionName.TARGETING_TACTICS,
    SectionName.TOPICS,
)

REGULAR_STATISTIC_SECTIONS = (
    SectionName.CREATIVES,
    SectionName.VIDEOS,
)

COLUMN_SET_REGULAR = (
    "Impressions",
    "Views",
    "View Rate",
    "Clicks",
    "CTR",
    "Video played to: 25%",
    "Video played to: 50%",
    "Video played to: 75%",
    "Video played to: 100%",
)

COLUMN_SET_WITH_CTA = (
    "Impressions",
    "Views",
    "View Rate",
    "Clicks",
    "Call-to-Action overlay",
    "Website",
    "App Store",
    "Cards",
    "End cap",
    "CTR",
    "Video played to: 25%",
    "Video played to: 50%",
    "Video played to: 75%",
    "Video played to: 100%",
)

COLUMN_SET_BY_SECTION_NAME = {
    **{section_name: COLUMN_SET_REGULAR for section_name in REGULAR_STATISTIC_SECTIONS},
    **{section_name: COLUMN_SET_WITH_CTA for section_name in SECTIONS_WITH_CTA},
}


class DashboardWeeklyReportAPITestCase(ExtendedAPITestCase):
    def _get_url(self, account_creation_id):
        return reverse(
            Name.Dashboard.PERFORMANCE_EXPORT_WEEKLY_REPORT,
            [RootNamespace.AW_CREATION, Namespace.DASHBOARD],
            args=(account_creation_id,)
        )

    def _request(self, account_creation_id):
        url = self._get_url(account_creation_id)
        return self.client.post(url, "{}", content_type="application/json")

    def test_no_demo_data(self):
        self.create_test_user()
        account = Account.objects.create(id=next(int_iterator))
        campaign_name = "Test campaign"
        Campaign.objects.create(name=campaign_name)

        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        self.assertTrue(is_report_empty(sheet))

    @generic_test([
        (section, (section,), dict())
        for section in SECTIONS_WITH_CTA
    ])
    def test_column_set(self, section):
        shared_columns = COLUMN_SET_BY_SECTION_NAME.get(section)
        self.create_test_user()
        account = Account.objects.create(id=next(int_iterator))

        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        row_index = get_section_start_row(sheet, section)
        title_values = tuple(cell.value for cell in sheet[row_index][1:])
        self.assertEqual(title_values, (section,) + shared_columns)

    def test_video_section(self):
        any_date_1 = date(2018, 1, 1)
        any_date_2 = any_date_1 + timedelta(days=1)
        today = max(any_date_1, any_date_2)
        self.create_test_user()
        any_video = SingleDatabaseApiConnectorPatcher.get_video_list()["items"][0]
        video_title = any_video["title"]
        video_id = any_video["id"]
        account = Account.objects.create(id=next(int_iterator))
        campaign = Campaign.objects.create(account=account)
        ad_group = AdGroup.objects.create(campaign=campaign)
        impressions = (2, 3)
        YTVideoStatistic.objects.create(date=any_date_1, ad_group=ad_group, yt_id=video_id, impressions=impressions[0])
        YTVideoStatistic.objects.create(date=any_date_2, ad_group=ad_group, yt_id=video_id, impressions=impressions[1])

        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }
        with patch_now(today), \
             self.patch_user_settings(**user_settings), \
             patch("aw_reporting.excel_reports_dashboard.SingleDatabaseApiConnector",
                   new=SingleDatabaseApiConnectorPatcher):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        row_index = get_section_start_row(sheet, SectionName.VIDEOS)
        data_row = sheet[row_index + 1]
        self.assertEqual(data_row[1].value, video_title)
        self.assertEqual(data_row[2].value, sum(impressions))

    def test_creatives_section(self):
        any_date_1 = date(2018, 1, 1)
        any_date_2 = any_date_1 + timedelta(days=1)
        today = max(any_date_1, any_date_2)
        self.create_test_user()
        any_video = SingleDatabaseApiConnectorPatcher.get_video_list()["items"][0]
        video_title = any_video["title"]
        video_id = any_video["id"]
        account = Account.objects.create(id=next(int_iterator))
        campaign = Campaign.objects.create(account=account)
        ad_group = AdGroup.objects.create(campaign=campaign)
        impressions = (2, 3)
        creative = VideoCreative.objects.create(id=video_id)
        common = dict(
            ad_group=ad_group,
            creative=creative,
        )
        VideoCreativeStatistic.objects.create(date=any_date_1, impressions=impressions[0], **common)
        VideoCreativeStatistic.objects.create(date=any_date_2, impressions=impressions[1], **common)

        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }
        with patch_now(today), \
             self.patch_user_settings(**user_settings), \
             patch("aw_reporting.excel_reports_dashboard.SingleDatabaseApiConnector",
                   new=SingleDatabaseApiConnectorPatcher):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        row_index = get_section_start_row(sheet, SectionName.CREATIVES)
        data_row = sheet[row_index + 1]
        self.assertEqual(data_row[1].value, video_title)
        self.assertEqual(data_row[2].value, sum(impressions))

    def test_ages_section(self):
        any_date_1 = date(2018, 1, 1)
        any_date_2 = any_date_1 + timedelta(days=1)
        today = max(any_date_1, any_date_2)
        self.create_test_user()
        account = Account.objects.create(id=next(int_iterator))
        campaign = Campaign.objects.create(account=account)
        ad_group = AdGroup.objects.create(campaign=campaign)
        age_range_id = AgeRange.AGE_45_54
        common = dict(
            age_range_id=age_range_id,
            ad_group=ad_group,
        )
        age_range_name = age_range_str(age_range_id)

        impressions = (2, 3)
        AgeRangeStatistic.objects.create(date=any_date_1, impressions=impressions[0], **common)
        AgeRangeStatistic.objects.create(date=any_date_2, impressions=impressions[1], **common)

        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }
        with patch_now(today), \
             self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        row_index = get_section_start_row(sheet, SectionName.AGES)
        data_row = sheet[row_index + 1]
        self.assertEqual(data_row[1].value, age_range_name)
        self.assertEqual(data_row[2].value, sum(impressions))

    def test_genders_section(self):
        any_date_1 = date(2018, 1, 1)
        any_date_2 = any_date_1 + timedelta(days=1)
        today = max(any_date_1, any_date_2)
        self.create_test_user()
        account = Account.objects.create(id=next(int_iterator))
        campaign = Campaign.objects.create(account=account)
        ad_group = AdGroup.objects.create(campaign=campaign)
        gender_id = Gender.MALE
        common = dict(
            gender_id=gender_id,
            ad_group=ad_group,
        )
        gender_name = gender_str(gender_id)

        impressions = (2, 3)
        GenderStatistic.objects.create(date=any_date_1, impressions=impressions[0], **common)
        GenderStatistic.objects.create(date=any_date_2, impressions=impressions[1], **common)

        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }
        with patch_now(today), \
             self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        row_index = get_section_start_row(sheet, SectionName.GENDERS)
        data_row = sheet[row_index + 1]
        self.assertEqual(data_row[1].value, gender_name)
        self.assertEqual(data_row[2].value, sum(impressions))

    def test_targeting_section(self):
        today = any_date = date(2018, 1, 1)
        self.create_test_user()
        account = Account.objects.create(id=next(int_iterator))
        campaign = Campaign.objects.create(account=account)
        ad_group = AdGroup.objects.create(campaign=campaign)

        common = dict(
            ad_group=ad_group,
            date=any_date,
        )
        TopicStatistic.objects.create(topic=Topic.objects.create(), impressions=2, **common)
        AudienceStatistic.objects.create(audience=Audience.objects.create(), impressions=3, **common)
        KeywordStatistic.objects.create(keyword="1", impressions=4, **common)
        YTChannelStatistic.objects.create(yt_id="1", impressions=5, **common)
        YTVideoStatistic.objects.create(yt_id="1", impressions=6, **common)

        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }
        with patch_now(today), \
             self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        start_row_index = get_section_start_row(sheet, SectionName.TARGETING_TACTICS)
        rows = [
            ("Topics", 2),
            ("Interests", 3),
            ("Keywords", 4),
            ("Channels", 5),
            ("Videos", 6),
        ]
        for index, data in enumerate(rows):
            name, value = data
            row_index = start_row_index + index + 1
            self.assertEqual(sheet[row_index][1].value, name, name)
            self.assertEqual(sheet[row_index][2].value, value, name)

    def test_budget(self):
        self.create_test_user()
        opportunity = Opportunity.objects.create(budget=123.45678)
        placement = OpPlacement.objects.create(opportunity=opportunity)
        account = Account.objects.create(id=next(int_iterator))
        Campaign.objects.create(account=account, salesforce_placement=placement)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }

        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        header_value = get_title_cell(sheet).value
        self.assertIn("Client Budget: ${}".format(opportunity.budget), header_value)

    def test_contracted_rate(self):
        self.create_test_user()
        opportunity = Opportunity.objects.create(contracted_cpm=.2, contracted_cpv=.5)
        placement = OpPlacement.objects.create(opportunity=opportunity)
        account = Account.objects.create(id=next(int_iterator))
        Campaign.objects.create(account=account, salesforce_placement=placement)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }

        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        header_value = get_title_cell(sheet).value
        expected_rates = "Contracted Rates: CPV ${} / CPM ${}".format(opportunity.contracted_cpv,
                                                                      opportunity.contracted_cpm)
        self.assertIn(expected_rates, header_value)

    def test_contracted_units(self):
        self.create_test_user()
        opportunity = Opportunity.objects.create(video_views=1234, impressions=6432)
        placement = OpPlacement.objects.create(opportunity=opportunity)
        account = Account.objects.create(id=next(int_iterator))
        Campaign.objects.create(account=account, salesforce_placement=placement)
        user_settings = {
            UserSettingsKey.VISIBLE_ALL_ACCOUNTS: True
        }

        with self.patch_user_settings(**user_settings):
            response = self._request(account.account_creation.id)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        header_value = get_title_cell(sheet).value
        expected_units = "Contracted Units: ordered CPV units = {} views / ordered CPM units = {} impressions".format(
            opportunity.video_views,
            opportunity.impressions,
        )
        self.assertIn(expected_units, header_value)

    @generic_test([
        (section, (section,), dict())
        for section in SECTIONS_WITH_CTA
    ])
    def test_demo_account_cta(self, section):
        shared_columns = COLUMN_SET_BY_SECTION_NAME.get(section)
        self.create_test_user()

        user_settings = {
            UserSettingsKey.DEMO_ACCOUNT_VISIBLE: True
        }
        with self.patch_user_settings(**user_settings):
            response = self._request(DEMO_ACCOUNT_ID)
        self.assertEqual(response.status_code, HTTP_200_OK)
        sheet = get_sheet_from_response(response)
        row_index = get_section_start_row(sheet, section)
        title_values = tuple(cell.value for cell in sheet[row_index][1:])
        self.assertEqual(title_values, (section,) + shared_columns)


def get_sheet_from_response(response):
    single_sheet_index = 0
    f = io.BytesIO(response.content)
    book = load_workbook(f)
    return book.worksheets[single_sheet_index]


TITLE_COLUMN = 1
TOTAL_NAME = "Total"
FIRST_SECTION_ROW_NUMBER = 13


def get_title_cell(sheet):
    return sheet[5][1]


def is_title_empty(sheet):
    title_cell = get_title_cell(sheet)
    return re.match(r"^Campaign: (None)?\n.*", title_cell.value) is not None


def is_demo_report(sheet):
    title_cell = get_title_cell(sheet)
    return re.match(r"^Campaign: Demo\n.*", title_cell.value) is not None and not is_title_empty(sheet)


def is_report_empty(sheet):
    return is_title_empty(sheet) and are_all_sections_empty(sheet)


def are_all_sections_empty(sheet):
    total_row = ("Total", 0)
    targeting_rows = ["Topics", "Interests", "Keywords", "Channels", "Videos"]
    section_names = (
        (SectionName.PLACEMENT, [total_row]),
        (SectionName.VIDEOS, [total_row]),
        (SectionName.CREATIVES, [total_row]),
        (SectionName.AD_GROUPS, []),
        (SectionName.INTERESTS, []),
        (SectionName.TOPICS, []),
        (SectionName.AGES, [total_row]),
        (SectionName.GENDERS, [total_row]),
        (SectionName.TARGETING_TACTICS, [(name, 0) for name in targeting_rows]),

        (SectionName.KEYWORDS, []),
        (SectionName.DEVICES, [(FOOTER_ANNOTATION, None)]),
    )
    return all([
        is_section_empty(sheet, section_name, static_rows)
        for section_name, static_rows in section_names
    ])


def is_static_row_empty(sheet, row_data, row):
    name, default_value = row_data
    return sheet[row][TITLE_COLUMN].value == name \
           and sheet[row][TITLE_COLUMN + 1].value == default_value


def is_section_empty(sheet, section_name, static_rows):
    section_row_number = get_section_start_row(sheet, section_name)
    static_rows_are_empty = all([
        is_static_row_empty(sheet, name, section_row_number + index + 1)
        for index, name in enumerate(static_rows)
    ])

    next_row_number = section_row_number + len(static_rows) + 1
    return static_rows_are_empty \
           and sheet[next_row_number][TITLE_COLUMN].value is None


def is_section_header(cell):
    return cell.alignment.horizontal == "center"


def get_section_start_row(sheet, section_name):
    data_rows_numbers = range(FIRST_SECTION_ROW_NUMBER, sheet.max_row)
    for row_number in data_rows_numbers:
        cell = sheet[row_number][TITLE_COLUMN]
        if is_section_header(cell) and cell.value == section_name:
            return row_number
    raise ValueError
